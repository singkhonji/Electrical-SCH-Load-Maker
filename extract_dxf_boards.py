"""
extract_dxf_boards.py
Extract every DB / panel from a DXF electrical schematic and write
each board to its own Excel sheet.

Usage:
    python extract_dxf_boards.py [input.dxf] [output.xlsx]
"""

import re
import os
import sys
import ezdxf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE      = os.path.dirname(os.path.abspath(__file__))
DXF_FILE  = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE, "input",  "D1-EMSB-BS1.dxf")
XLSX_FILE = sys.argv[2] if len(sys.argv) > 2 else os.path.join(BASE, "output", "D1-EMSB-BS1_Boards.xlsx")

# ═══════════════════════════════════════════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════════════════════════════════════════
def fill(h):
    return PatternFill("solid", fgColor="FF" + h.upper())

F_NAVY   = fill("1F3864")
F_BLUE   = fill("2E75B6")
F_LBLUE  = fill("D6E4F0")
F_WHITE  = fill("FFFFFF")
F_PEACH  = fill("FCE4D6")
F_GREEN  = fill("E2EFDA")
F_LGRAY  = fill("F2F2F2")
F_YELLOW = fill("FFF2CC")

def fnt(bold=False, sz=9, col="000000"):
    return Font(name="Calibri", bold=bold, size=sz, color="FF" + col.upper())

FT_TITLE = fnt(True,  13, "FFFFFF")
FT_SUB   = fnt(False,  9, "FFFFFF")
FT_HDR   = fnt(True,   9, "FFFFFF")
FT_LBL   = fnt(True,   9)
FT_BODY  = fnt(False,  9)
FT_BOLD  = fnt(True,   9)
FT_GREY  = fnt(False,  9, "7F7F7F")
FT_GBOLD = fnt(True,   9, "7F7F7F")

AC  = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AR  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
_t  = Side(style="thin", color="FF000000")
BD  = Border(left=_t, right=_t, top=_t, bottom=_t)
N   = 10  # columns A–J


def wc(ws, r, c, v, ft=FT_BODY, fg=F_WHITE, al=AC):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = ft; cell.fill = fg; cell.alignment = al; cell.border = BD
    return cell


def mw(ws, r1, c1, r2, c2, v, ft=FT_BODY, fg=F_WHITE, al=AC):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=v)
    cell.font = ft; cell.fill = fg; cell.alignment = al; cell.border = BD
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(row=rr, column=cc).border = BD
    return cell


# ═══════════════════════════════════════════════════════════════════════════════
# DXF LOADING
# ═══════════════════════════════════════════════════════════════════════════════
def load_ents(path):
    """Return list of {'x', 'y', 'text'} for all TEXT/MTEXT in modelspace."""
    doc = ezdxf.readfile(path)
    out = []
    for e in doc.modelspace():
        if e.dxftype() == "TEXT":
            t = e.dxf.text.strip()
        elif e.dxftype() == "MTEXT":
            t = e.plain_text().strip()
        else:
            continue
        if t:
            out.append({"x": e.dxf.insert.x, "y": e.dxf.insert.y, "text": t})
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# BOARD DETECTION
# ═══════════════════════════════════════════════════════════════════════════════
# Matches board designations like D1-SB-COM-FOH-BS1-A, D1-DB-MDF3SGB-A etc.
# Must start with D1-, contain at least 3 dash-separated parts, end with a
# single letter suffix (the phase / board letter).
BOARD_RE = re.compile(r"^D1-[A-Z0-9]+(?:-[A-Z0-9]+){2,}-[A-Z]$")


def find_boards(ents):
    """
    Find unique board title entities.
    Also handles combined labels like 'D1-DB-MDF3SGB-A/D1-DB-MDF2SGB-A'
    by splitting on '/'.
    Returns: list of {'name', 'x', 'y'}
    """
    seen   = {}
    boards = []
    for e in ents:
        for part in e["text"].split("/"):
            t = part.strip()
            if BOARD_RE.match(t) and t not in seen:
                seen[t] = True
                boards.append({"name": t, "x": e["x"], "y": e["y"]})
    return sorted(boards, key=lambda b: (-b["y"], b["x"]))


def cluster_by_y(boards, tol=5000):
    """Group boards into horizontal bands (Y levels)."""
    levels = []
    for b in sorted(boards, key=lambda b: -b["y"]):
        placed = False
        for lv in levels:
            if abs(lv["y_center"] - b["y"]) <= tol:
                lv["boards"].append(b)
                lv["y_center"] = sum(bb["y"] for bb in lv["boards"]) / len(lv["boards"])
                placed = True
                break
        if not placed:
            levels.append({"y_center": b["y"], "boards": [b]})
    for lv in levels:
        lv["boards"].sort(key=lambda b: b["x"])
    return sorted(levels, key=lambda l: -l["y_center"])


# ═══════════════════════════════════════════════════════════════════════════════
# SPATIAL HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def ents_in_box(ents, x_lo, x_hi, y_lo, y_hi):
    return [e for e in ents if x_lo <= e["x"] <= x_hi and y_lo <= e["y"] <= y_hi]


def nearest(ents, x, y, x_tol=15000, y_tol=3000):
    m = [e for e in ents if abs(e["x"] - x) <= x_tol and abs(e["y"] - y) <= y_tol]
    if not m:
        return ""
    return min(m, key=lambda e: abs(e["x"] - x) + abs(e["y"] - y))["text"]


def y_of(ents, marker):
    """Return Y coordinate of the first entity whose text contains 'marker'."""
    for e in sorted(ents, key=lambda e: -e["y"]):
        if marker.lower() in e["text"].lower():
            return e["y"]
    return None


def band(ents, y_ref, tol=2500):
    """All entities within ±tol of y_ref, sorted by X."""
    return sorted([e for e in ents if abs(e["y"] - y_ref) <= tol],
                  key=lambda e: e["x"])


def x_boundaries_from_circuits(all_ents, level_boards, y_lo, y_hi):
    """
    Determine X-split points between boards at this level by finding natural
    gaps in circuit-number (RYB) X positions.  Falls back to board-title
    midpoints if no circuits are found.
    Returns sorted list of (x_lo, x_hi) tuples, one per board (same order
    as level_boards sorted by X).
    """
    bs = sorted(level_boards, key=lambda b: b["x"])
    n  = len(bs)

    # Collect all RYB circuit X coords in the level's Y band
    y_ckt = y_of(ents_in_box(all_ents, -1e9, 1e9, y_lo, y_hi), "CIRCUIT NO")
    ckt_xs = []
    if y_ckt:
        for e in band(ents_in_box(all_ents, -1e9, 1e9, y_lo, y_hi), y_ckt, 3000):
            if (re.search(r"(?:P-\d+)?[0-9]*RYB", e["text"], re.I)
                    and "CIRCUIT" not in e["text"].upper()):
                ckt_xs.append(e["x"])

    if len(ckt_xs) >= n and n > 1:
        ckt_xs.sort()
        # Find the (n-1) largest X gaps → natural board boundaries
        gaps = sorted(
            [(ckt_xs[i + 1] - ckt_xs[i], (ckt_xs[i] + ckt_xs[i + 1]) / 2)
             for i in range(len(ckt_xs) - 1)],
            key=lambda g: -g[0],
        )
        splits = sorted(g[1] for g in gaps[: n - 1])
    else:
        # Fallback: midpoints between board title X positions
        splits = [(bs[i]["x"] + bs[i + 1]["x"]) / 2 for i in range(n - 1)]

    # Build (x_lo, x_hi) pairs
    margin = 50000
    xs = [-1e9] + splits + [1e9]
    return [(xs[i], xs[i + 1]) for i in range(n)]


# ═══════════════════════════════════════════════════════════════════════════════
# BOARD REGION  (now uses circuit-gap–derived X boundaries)
# ═══════════════════════════════════════════════════════════════════════════════
def board_ents(all_ents, board, x_lo, x_hi, y_lo, y_hi):
    return ents_in_box(all_ents, x_lo, x_hi, y_lo, y_hi)


# ═══════════════════════════════════════════════════════════════════════════════
# PARSE ONE BOARD
# ═══════════════════════════════════════════════════════════════════════════════
_SKIP = {"ROOM", "NAME", "CAPACITY", "CIRCUIT", "INSTALL",
         "CABLE", "XLPE", "CONDUIT", "TRUNKING", "LADDER",
         "REASON:", "FOR TENDER", "2025.", "TOTAL:"}

def _skip(text):
    tu = text.upper()
    return (
        any(s in tu for s in _SKIP)
        or tu in ("FS", "ST", "L1", "L2", "L3", "L", "N", "PE",
                  "SPD", "ZCT", "ELR", "BMS", "PM")
        or re.match(r"^\d+A$", text)
        or len(text) <= 1
    )


def parse_board(board, ents):
    name = board["name"]

    # ── Header fields ──────────────────────────────────────────────────────────
    for_desc = next((e["text"] for e in ents
                     if e["text"].upper().startswith("FOR ")), "")
    supply_raw = next((e["text"] for e in ents
                       if e["text"].upper().startswith("FROM ")), "")
    supply = re.sub(r"^FROM\s+", "", supply_raw, flags=re.I).strip()

    incomer_cb = next(
        (e["text"].replace("\n", " ")
         for e in ents
         if re.search(r"\d+AT.*MCCB|\d+AT.*TPN|TPN\s*MCCB.*L,S,I", e["text"], re.I)),
        "",
    )

    busbar = next(
        (e["text"] for e in ents if "RATED 4P INSULATED" in e["text"].upper()),
        "",
    )

    incomer_cable = next(
        (e["text"].replace("\n", " ")
         for e in ents
         if re.search(r"(?:70|35)mm", e["text"]) and "XLPE" in e["text"].upper()),
        "",
    )

    # ── Circuit numbers ────────────────────────────────────────────────────────
    # Find the Y of the circuit-number row.
    # Primary: look for the "CIRCUIT NO:" header row.
    # Fallback: detect RYB-pattern entities directly.
    y_ckt  = y_of(ents, "CIRCUIT NO")
    y_room = y_of(ents, "ROOM / EQUIPMENT")

    if y_ckt is None:
        # "CIRCUIT NO:" header is outside our X bounding box — infer Y from
        # the RYB entities themselves.
        ryb_scan = [e for e in ents
                    if re.search(r"^(?:P-\d+)?[0-9]*RYB$", e["text"].strip(), re.I)]
        if ryb_scan:
            y_ckt = sorted(ryb_scan, key=lambda e: e["y"])[len(ryb_scan) // 2]["y"]

    circuits = []
    if y_ckt:
        for e in band(ents, y_ckt, 3000):
            t = e["text"].strip()
            if re.search(r"(?:P-\d+)?[0-9]*RYB", t, re.I) and "CIRCUIT" not in t.upper():
                circuits.append({"x": e["x"], "circuit_no": t,
                                  "desc": "", "cb": "", "cable": ""})

    # De-duplicate by X proximity
    circuits_dedup = []
    for c in sorted(circuits, key=lambda c: c["x"]):
        if not any(abs(d["x"] - c["x"]) < 1000 for d in circuits_dedup):
            circuits_dedup.append(c)
    circuits = circuits_dedup

    if not circuits:
        return {"name": name, "for_desc": for_desc, "supply": supply,
                "incomer_cb": incomer_cb, "busbar": busbar,
                "incomer_cable": incomer_cable, "circuits": []}

    # ── Descriptions (look 2000–15000 units above circuit row) ────────────────
    y_desc_candidates = []
    if y_room:
        y_desc_candidates.append(y_room)
    if y_ckt:
        for dy in (2000, 3000, 4000, 6000, 8000, 10000, 12000, 15000):
            y_desc_candidates.append(y_ckt + dy)

    for circ in circuits:
        for y_c in y_desc_candidates:
            nearby = [e for e in ents
                      if abs(e["x"] - circ["x"]) <= 6000
                      and abs(e["y"] - y_c) <= 2000
                      and not _skip(e["text"])]
            if nearby:
                best = min(nearby, key=lambda e: abs(e["x"] - circ["x"]))
                circ["desc"] = best["text"]
                if circ["desc"] != "SPARE":
                    break

    # ── CB ratings ─────────────────────────────────────────────────────────────
    # Find incomer CB X so we don't accidentally pick it up as a circuit CB
    incomer_x = next(
        (e["x"] for e in ents
         if re.search(r"\d+AT.*MCCB", e["text"], re.I)),
        None,
    )
    cb_ents = [e for e in ents if re.search(r"\d+A.*(?:MCCB|MCB|RCCB)", e["text"], re.I)]

    for circ in circuits:
        nearby = [e for e in cb_ents
                  if abs(e["x"] - circ["x"]) <= 5000
                  and (incomer_x is None or abs(e["x"] - incomer_x) > 2000)]
        if nearby:
            circ["cb"] = min(nearby, key=lambda e: abs(e["x"] - circ["x"]))["text"].replace("\n", " ")

    # ── Cable data ─────────────────────────────────────────────────────────────
    cable_ents = [e for e in ents
                  if re.search(r"\d+mm.*CU", e["text"], re.I)
                  and "XLPE/LSZH CU\nOR" not in e["text"]]  # exclude header row

    for circ in circuits:
        nearby = [e for e in cable_ents if abs(e["x"] - circ["x"]) <= 5000]
        if nearby:
            circ["cable"] = min(nearby, key=lambda e: abs(e["x"] - circ["x"]))["text"].replace("\n", " ")

    return {
        "name":           name,
        "for_desc":       for_desc,
        "supply":         supply,
        "incomer_cb":     incomer_cb,
        "busbar":         busbar,
        "incomer_cable":  incomer_cable,
        "circuits":       circuits,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# WRITE EXCEL SHEET
# ═══════════════════════════════════════════════════════════════════════════════
HDR_LABELS = [
    "Circuit No.", "Description / Load Name", "CB Rating (A)", "CB Type",
    "No. of Poles", "Breaking Cap.", "Cable Size", "Cable Route", "Status", "Remarks",
]
COL_W = {"A": 16, "B": 38, "C": 14, "D": 14,
         "E": 9,  "F": 12, "G": 30, "H": 22, "I": 9, "J": 14}


def _parse_cb(cb_txt):
    """Split a CB text block into (rating_A, type, poles, breaking_cap)."""
    parts = re.findall(r"\S+", cb_txt)
    cb_a   = next((p for p in parts if re.match(r"^\d+A$", p)), "-")
    cb_t   = ("MCCB" if "MCCB" in cb_txt
               else "MCB"  if "MCB"  in cb_txt
               else "RCCB" if "RCCB" in cb_txt
               else "-")
    poles  = next((p for p in parts
                   if re.match(r"^(?:[1-4]P\+?N?|TPN|SPN)$", p, re.I)), "-")
    brk    = next((p for p in parts if re.match(r"^\d+k[Aa]$", p, re.I)), "-")
    return cb_a, cb_t, poles, brk


def write_sheet(wb, data):
    sname = re.sub(r"[\\/*?:\[\]]", "-", data["name"])[:31]
    ws    = wb.create_sheet(title=sname)

    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w

    # Row 1 — Title
    mw(ws, 1, 1, 1, N, f"LOAD SCHEDULE — {data['name']}",
       ft=FT_TITLE, fg=F_NAVY)
    ws.row_dimensions[1].height = 22

    # Row 2 — Subtitle
    mw(ws, 2, 1, 2, N, data["for_desc"] or data["name"],
       ft=FT_SUB, fg=F_BLUE)
    ws.row_dimensions[2].height = 16

    # Row 3 — Board / Source
    wc(ws, 3, 1, "Board Designation:", ft=FT_LBL, fg=F_LBLUE, al=AL)
    mw(ws, 3, 2, 3, 5, data["name"],  ft=FT_BODY, fg=F_WHITE, al=AL)
    wc(ws, 3, 6, "Supply Source:",     ft=FT_LBL, fg=F_LBLUE, al=AL)
    mw(ws, 3, 7, 3, N,  data["supply"],ft=FT_BODY, fg=F_WHITE, al=AL)
    ws.row_dimensions[3].height = 16

    # Row 4 — Incomer CB / Busbar
    wc(ws, 4, 1, "Incomer CB:",        ft=FT_LBL, fg=F_LBLUE, al=AL)
    mw(ws, 4, 2, 4, 5, data["incomer_cb"],  ft=FT_BODY, fg=F_WHITE, al=AL)
    wc(ws, 4, 6, "Busbar Rating:",      ft=FT_LBL, fg=F_LBLUE, al=AL)
    mw(ws, 4, 7, 4, N, data["busbar"],  ft=FT_BODY, fg=F_WHITE, al=AL)
    ws.row_dimensions[4].height = 16

    # Row 5 — Incomer Cable
    wc(ws, 5, 1, "Incomer Cable:",      ft=FT_LBL, fg=F_LBLUE, al=AL)
    mw(ws, 5, 2, 5, N, data["incomer_cable"], ft=FT_BODY, fg=F_WHITE, al=AL)
    ws.row_dimensions[5].height = 16

    # Row 6 — spacer
    ws.row_dimensions[6].height = 6

    # Row 7 — Section bar
    mw(ws, 7, 1, 7, N, "OUTGOING CIRCUITS", ft=FT_HDR, fg=F_BLUE)
    ws.row_dimensions[7].height = 16

    # Row 8 — Column headers
    for i, h in enumerate(HDR_LABELS, 1):
        wc(ws, 8, i, h, ft=FT_HDR, fg=F_NAVY)
    ws.row_dimensions[8].height = 30

    # Circuit rows
    circuits = data["circuits"]
    if not circuits:
        mw(ws, 9, 1, 9, N,
           "No circuits extracted — verify DXF layout manually.",
           ft=FT_BODY, fg=F_YELLOW, al=AL)
        ws.row_dimensions[9].height = 18
        ws.freeze_panes = "A9"
        return

    for i, circ in enumerate(circuits):
        r     = 9 + i
        no    = circ["circuit_no"]
        desc  = circ.get("desc", "") or ""
        spare = "SPARE" in no.upper() or desc.upper() == "SPARE"

        fg_r  = F_GREEN if spare else (F_LBLUE if i % 2 == 0 else F_WHITE)
        ft_no = FT_GBOLD if spare else FT_BOLD
        ft_b  = FT_GREY  if spare else FT_BODY

        cb_a, cb_t, poles, brk = _parse_cb(circ.get("cb", ""))

        wc(ws, r, 1, no,                          ft=ft_no, fg=fg_r, al=AC)
        wc(ws, r, 2, desc or ("-" if spare else "?"), ft=ft_b, fg=fg_r, al=AL)
        wc(ws, r, 3, cb_a,                         ft=ft_b, fg=fg_r, al=AC)
        wc(ws, r, 4, cb_t,                         ft=ft_b, fg=fg_r, al=AC)
        wc(ws, r, 5, poles,                        ft=ft_b, fg=fg_r, al=AC)
        wc(ws, r, 6, brk,                          ft=ft_b, fg=fg_r, al=AC)
        wc(ws, r, 7, circ.get("cable", "") or "-", ft=ft_b, fg=fg_r, al=AL)
        wc(ws, r, 8, "-",                          ft=ft_b, fg=fg_r, al=AL)
        wc(ws, r, 9, "SPARE" if spare else "ACTIVE", ft=ft_b, fg=fg_r, al=AC)
        wc(ws, r, 10, "",                          ft=ft_b, fg=fg_r, al=AC)
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A9"


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    print(f"Reading: {DXF_FILE}")
    ents = load_ents(DXF_FILE)
    print(f"  {len(ents)} text entities loaded")

    boards = find_boards(ents)
    if not boards:
        print("  ERROR: No board titles detected. Exiting.")
        return

    print(f"  {len(boards)} boards detected:")
    for b in boards:
        print(f"    {b['name']:40s}  X={b['x']:9.0f}  Y={b['y']:9.0f}")

    levels = cluster_by_y(boards)
    print(f"\n  {len(levels)} Y-levels:")
    for lv in levels:
        print(f"    Y≈{lv['y_center']:.0f} : {[b['name'] for b in lv['boards']]}")

    wb = Workbook()
    wb.remove(wb.active)   # discard default empty sheet

    for li, level in enumerate(levels):
        level_y       = level["y_center"]
        lower_level_y = levels[li + 1]["y_center"] if li < len(levels) - 1 else None

        y_hi_lv = level_y + 5000
        y_lo_lv = lower_level_y + 2000 if lower_level_y is not None else level_y - 65000

        # X boundaries derived from actual circuit gap positions
        bs_sorted = sorted(level["boards"], key=lambda b: b["x"])
        x_spans   = x_boundaries_from_circuits(ents, level["boards"], y_lo_lv, y_hi_lv)

        for board, (x_lo, x_hi) in zip(bs_sorted, x_spans):
            bents = board_ents(ents, board, x_lo, x_hi, y_lo_lv, y_hi_lv)
            data  = parse_board(board, bents)

            nckt = len(data["circuits"])
            nact = sum(1 for c in data["circuits"]
                       if "SPARE" not in c["circuit_no"].upper())
            print(f"  Sheet: {board['name']:40s}  "
                  f"ents={len(bents):4d}  circuits={nckt:2d}  active={nact}")

            write_sheet(wb, data)

    os.makedirs(os.path.dirname(XLSX_FILE), exist_ok=True)
    wb.save(XLSX_FILE)
    print(f"\nSaved → {XLSX_FILE}")


if __name__ == "__main__":
    main()
