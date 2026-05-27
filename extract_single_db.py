"""
extract_single_db.py
Extracts a single Distribution Board load schedule from a DXF file
and writes a formatted Excel load schedule matching the reference template.

Usage:  python extract_single_db.py input/D1-DB-COM-FOHRA-A.dxf
Output: output/<board_name>_LoadSchedule.xlsx
"""

import os, re, sys
from collections import Counter
import ezdxf
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def load_ents(path):
    doc = ezdxf.readfile(path)
    msp = doc.modelspace()
    ents = []
    for e in msp:
        if e.dxftype() == 'TEXT':
            ents.append({'x': round(e.dxf.insert.x), 'y': round(e.dxf.insert.y),
                         'text': (e.dxf.text or '').strip()})
        elif e.dxftype() == 'MTEXT':
            ents.append({'x': round(e.dxf.insert.x), 'y': round(e.dxf.insert.y),
                         'text': (e.plain_text() or '').strip()})
    return [e for e in ents if e['text']]


def near_x(candidates, x, tol=3000):
    best = min(candidates, key=lambda e: abs(e['x'] - x), default=None)
    if best and abs(best['x'] - x) <= tol:
        return best['text']
    return ''


def band(ents, y_center, half=2000):
    return [e for e in ents if abs(e['y'] - y_center) <= half]


def first_match(ents, pattern, flags=re.I):
    for e in sorted(ents, key=lambda e: -e['y']):
        if re.search(pattern, e['text'], flags):
            return e['text']
    return ''


def y_of(ents, keyword):
    for e in sorted(ents, key=lambda e: -e['y']):
        if keyword.lower() in e['text'].lower():
            return e['y']
    return None


def clean_text(t):
    """Normalise superscripts and whitespace from DXF text."""
    t = t.replace('\n', ' ').strip()
    t = re.sub(r'mm[2²]\^?', 'mm²', t)
    t = t.replace('2^', '²').replace('^', '²')
    t = re.sub(r'\s{2,}', ' ', t)
    return t


# ══════════════════════════════════════════════════════════════════════════════
# CIRCUIT NUMBER PARSING
# ══════════════════════════════════════════════════════════════════════════════

_FIRST_CKT = re.compile(r'^([A-Z][A-Z0-9]*)[-/](\d+)([RYB])$', re.I)
_CONT_CKT  = re.compile(r'^(\d+)([RYB])$', re.I)


def parse_circuit_ents(circuit_ents):
    results = []
    current_group = ''
    for e in sorted(circuit_ents, key=lambda e: e['x']):
        t = e['text'].strip()
        m = _FIRST_CKT.match(t)
        if m:
            current_group = m.group(1).upper()
            num, phase = int(m.group(2)), m.group(3).upper()
        else:
            m2 = _CONT_CKT.match(t)
            if m2:
                num, phase = int(m2.group(1)), m2.group(2).upper()
            else:
                continue
        full_id = f"{current_group}-{num}{phase}" if current_group else t
        results.append({'x': e['x'], 'y': e['y'],
                        'circuit_no': full_id,
                        'group': current_group,
                        'num': num, 'phase': phase})
    return results


# ══════════════════════════════════════════════════════════════════════════════
# CB / CABLE PARSING
# ══════════════════════════════════════════════════════════════════════════════

def parse_cb(text):
    """Return (rating, cb_type, poles, breaking) from breaker text."""
    t = clean_text(text)
    m = re.search(r'(\d+)\s*A[TF]?', t, re.I)
    rating = f"{m.group(1)}A" if m else ''
    if   'RCCB' in t.upper(): cb_type = 'RCCB'
    elif 'MCCB' in t.upper(): cb_type = 'MCCB'
    elif 'MCB'  in t.upper(): cb_type = 'MCB'
    else:                       cb_type = ''
    if   re.search(r'\bSPN\b|\b1P\b', t, re.I):  poles = '1P'
    elif re.search(r'\bTPN\b',        t, re.I):   poles = '3P+N'
    elif re.search(r'\b4P\b',         t, re.I):   poles = '4P+N'
    elif re.search(r'\b2P\b',         t, re.I):   poles = '2P'
    else:                                           poles = ''
    m2 = re.search(r'(\d+)\s*KA',     t, re.I)
    m3 = re.search(r'\((\d+)\s*mA\)', t, re.I)
    brk = (f"{m2.group(1)}kA" if m2 else f"{m3.group(1)}mA" if m3 else '')
    if re.search(r'[LSI],[LSI],[LSI]', t):
        cb_type += ' (L,S,I)'
    return rating, cb_type.strip(), poles, brk


def parse_cable(text):
    """Return (size, route) from cable description text."""
    t = clean_text(text)
    parts = re.split(r'\s+ON\s+', t, maxsplit=1, flags=re.I)
    size  = parts[0].strip()
    route = parts[1].strip().title() if len(parts) > 1 else ''
    return size, route


# ══════════════════════════════════════════════════════════════════════════════
# DATA EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_board(ents):
    # Board name: top-most entity matching board name pattern
    board_name = ''
    for e in sorted(ents, key=lambda e: -e['y']):
        if re.match(r'^D\d+-[A-Z]', e['text'], re.I):
            board_name = e['text']
            break

    for_desc    = first_match(ents, r'^FOR\s+',  re.I)
    from_supply = first_match(ents, r'^FROM\s+', re.I)

    # Incomer: highest-Y MCCB text
    mccb_ents = sorted(
        [e for e in ents if re.search(r'\d+A[TF]?\s*\n?\s*(?:TPN|SPN)?\s*MCCB', e['text'], re.I)],
        key=lambda e: -e['y'])
    incomer_raw = mccb_ents[0]['text'] if mccb_ents else ''
    inc_rating, inc_type, inc_poles, inc_brk = parse_cb(incomer_raw)

    busbar   = first_match(ents, r'BUSBAR', re.I)
    feed_raw = first_match(ents, r'^\d+\s*x\s*1C', re.I)
    feed_size, feed_route = parse_cable(feed_raw) if feed_raw else ('', '')

    # ── Key Y levels ──────────────────────────────────────────────────────────
    y_ckt_hdr  = y_of(ents, "CIRCUIT NO")
    y_kw_hdr   = y_of(ents, "CAPACITY")
    y_room_hdr = y_of(ents, "ROOM / EQUIPMENT")

    if y_ckt_hdr is None:
        raise ValueError("Cannot find 'CIRCUIT NO' in DXF")

    # Circuit entities
    ckt_ents = [e for e in band(ents, y_ckt_hdr, 500)
                if _FIRST_CKT.match(e['text']) or _CONT_CKT.match(e['text'])]
    circuits = parse_circuit_ents(ckt_ents)
    if not circuits:
        raise ValueError("No circuit entities found near CIRCUIT NO row")

    # Data bands
    kw_ents   = band(ents, y_kw_hdr + 50, 300) if y_kw_hdr else []
    desc_ents = band(ents, y_room_hdr + 300, 500) if y_room_hdr else []

    # Cable: largest Y-cluster of cable texts
    cable_all = [e for e in ents if re.search(r'\d+mm[²2^]', e['text'], re.I)]
    if cable_all:
        yr    = [round(e['y'] / 500) * 500 for e in cable_all]
        dom_y = Counter(yr).most_common(1)[0][0]
        cable_band = [e for e in cable_all if abs(round(e['y'] / 500) * 500 - dom_y) < 1]
    else:
        cable_band = []

    # Breakers
    brk_all   = [e for e in ents
                 if re.search(r'\d+A\s*(?:SPN|TPN|4P|2P)?\s*(?:MCB|MCCB|RCCB)', e['text'], re.I)]
    spn_ents  = [e for e in brk_all if re.search(r'\bSPN\b', e['text'], re.I)]
    rccb_ents = [e for e in brk_all if re.search(r'\bRCCB\b', e['text'], re.I)]

    # Accessories
    def find_kw(kw):
        matches = [e for e in ents if kw.lower() in e['text'].lower()]
        return matches[0]['text'] if matches else ''

    acc_spd_spec = first_match(ents, r'8/20', re.I)
    acc_spd_mcb  = ''
    spd_ents = [e for e in ents if e['text'].strip() == 'SPD']
    if spd_ents:
        spd_x = spd_ents[0]['x']
        nb = near_x([e for e in ents if re.search(r'^\d+A$', e['text'])], spd_x, 3000)
        acc_spd_mcb = nb

    acc_pm_cls = next((e['text'] for e in ents if 'CLASS 0.5' in e['text'].upper()), '')
    acc_lt     = first_match(ents, r'LIGHTING CONTROL|TIMER.*SENSOR', re.I)

    # ── Per-circuit data ──────────────────────────────────────────────────────
    for c in circuits:
        x = c['x']
        c['kw']        = near_x(kw_ents,    x, tol=2000)
        c['desc']      = near_x(desc_ents,  x, tol=5000)
        c['cb_raw']    = near_x(spn_ents,   x, tol=2000)
        c['cable_raw'] = near_x(cable_band, x, tol=2000)
        c['rccb_raw']  = near_x(rccb_ents,  x, tol=15000)

        if c['cb_raw']:
            c['cb_rating'], c['cb_type'], c['cb_poles'], c['cb_brk'] = parse_cb(c['cb_raw'])
        else:
            c['cb_rating'] = c['cb_type'] = c['cb_poles'] = c['cb_brk'] = ''

        if c['cable_raw']:
            c['cable_size'], c['cable_route'] = parse_cable(c['cable_raw'])
        else:
            c['cable_size'] = c['cable_route'] = ''

        if c['rccb_raw']:
            c['rccb_rating'], c['rccb_type'], c['rccb_poles'], c['rccb_brk'] = parse_cb(c['rccb_raw'])
        else:
            c['rccb_rating'] = c['rccb_type'] = c['rccb_poles'] = c['rccb_brk'] = ''

        c['status'] = 'SPARE' if 'SPARE' in c['desc'].upper() else 'ACTIVE'

    return {
        'name':        board_name,
        'for_desc':    for_desc,
        'from_supply': from_supply,
        'incomer': {'raw': incomer_raw, 'rating': inc_rating,
                    'type': inc_type, 'poles': inc_poles, 'brk': inc_brk},
        'busbar':      clean_text(busbar),
        'feed_size':   feed_size,
        'feed_route':  feed_route,
        'circuits':    circuits,
        'accessories': {
            'spd':     acc_spd_spec,
            'spd_mcb': acc_spd_mcb,
            'pm':      acc_pm_cls,
            'lt_ctrl': acc_lt,
            'elr':     find_kw('ELR'),
            'zct':     find_kw('ZCT'),
            'bms':     find_kw('BMS'),
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL STYLES  (matching reference template)
# ══════════════════════════════════════════════════════════════════════════════

def fill(hex6):
    return PatternFill("solid", fgColor="FF" + hex6.upper())

F_NAVY  = fill("1F3864")
F_BLUE  = fill("2E75B6")
F_LBLUE = fill("D6E4F0")
F_WHITE = fill("FFFFFF")
F_PEACH = fill("FCE4D6")
F_GREEN = fill("E2EFDA")
F_LGRAY = fill("F2F2F2")
F_YELL  = fill("FFF2CC")
F_RCCB  = fill("D9E1F2")

def fnt(bold=False, size=9, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size,
                color="FF" + color.upper(), italic=italic)

FT_TITLE = fnt(bold=True,  size=13, color="FFFFFF")
FT_SUB   = fnt(bold=False, size=9,  color="FFFFFF")
FT_HDR   = fnt(bold=True,  size=9,  color="FFFFFF")
FT_LABEL = fnt(bold=True,  size=9)
FT_BODY  = fnt(size=9)
FT_BOLD  = fnt(bold=True,  size=9)
FT_GREY  = fnt(size=9,  color="7F7F7F")
FT_GBOLD = fnt(bold=True, size=9, color="7F7F7F")
FT_RCCB  = fnt(bold=True, size=9, color="1F3864")

AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_thin = Side(style="thin", color="FF000000")
BD = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

N = 14

COL_W = {
    "A": 16, "B": 38, "C": 11, "D": 14,
    "E": 8,  "F": 10, "G": 11, "H": 9,
    "I": 10, "J": 9,  "K": 28, "L": 22,
    "M": 12, "N": 9,
}

HDR_LABELS = [
    "Circuit No.", "Description / Load Name",
    "CB Rating\n(A)", "CB Type",
    "No. of\nPoles", "Breaking\nCap.",
    "Connected\nLoad (kW)", "Demand\nFactor",
    "Max Demand\n(kW)", "Current\n(A)",
    "Cable Size", "Cable Route",
    "Remarks", "Status",
]


def w(ws, r, c, val, ft=None, fg=None, al=AC):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = ft or FT_BODY
    cell.fill      = fg or F_WHITE
    cell.alignment = al
    cell.border    = BD
    return cell


def mw(ws, r1, c1, r2, c2, val, ft=None, fg=None, al=AC):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    cell.font      = ft or FT_BODY
    cell.fill      = fg or F_WHITE
    cell.alignment = al
    cell.border    = BD
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(row=rr, column=cc).border = BD
    return cell


def section_bar(ws, r, text):
    mw(ws, r, 1, r, N, text, ft=FT_HDR, fg=F_BLUE, al=AC)
    ws.row_dimensions[r].height = 15.75


def col_headers(ws, r):
    for i, lbl in enumerate(HDR_LABELS, 1):
        w(ws, r, i, lbl, ft=FT_HDR, fg=F_NAVY, al=AC)
    ws.row_dimensions[r].height = 33.75


def circuit_row(ws, r, cno, desc, cb_a, cb_t, poles, brk,
                kw, cable_sz, cable_rt, remarks, status, fg_row):
    is_spare = (status == 'SPARE')
    ft_id  = FT_GBOLD if is_spare else FT_BOLD
    ft_dat = FT_GREY  if is_spare else FT_BODY
    ft_st  = FT_GBOLD if is_spare else FT_BOLD

    w(ws, r, 1,  cno,        ft=ft_id,  fg=fg_row, al=AC)
    w(ws, r, 2,  desc,       ft=ft_dat, fg=fg_row, al=AL)
    w(ws, r, 3,  cb_a,       ft=ft_dat, fg=fg_row, al=AC)
    w(ws, r, 4,  cb_t,       ft=ft_dat, fg=fg_row, al=AC)
    w(ws, r, 5,  poles,      ft=ft_dat, fg=fg_row, al=AC)
    w(ws, r, 6,  brk,        ft=ft_dat, fg=fg_row, al=AC)
    w(ws, r, 7,  kw or None, ft=ft_dat, fg=fg_row, al=AC)
    w(ws, r, 8,  None,       ft=ft_dat, fg=fg_row, al=AC)
    cell = ws.cell(row=r, column=9, value=f"=G{r}*H{r}")
    cell.font = ft_dat; cell.fill = fg_row; cell.alignment = AC; cell.border = BD
    w(ws, r, 10, None,       ft=ft_dat, fg=fg_row, al=AC)
    w(ws, r, 11, cable_sz,   ft=ft_dat, fg=fg_row, al=AL)
    w(ws, r, 12, cable_rt,   ft=ft_dat, fg=fg_row, al=AL)
    w(ws, r, 13, remarks,    ft=ft_dat, fg=fg_row, al=AC)
    w(ws, r, 14, status,     ft=ft_st,  fg=fg_row, al=AC)
    ws.row_dimensions[r].height = 27.75


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════════════

def write_excel(data, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data['name'][:31]
    ws.sheet_properties.tabColor = "2E75B6"

    for col, width in COL_W.items():
        ws.column_dimensions[col].width = width

    row = 1

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    mw(ws, row, 1, row, N,
       f"LOAD SCHEDULE — {data['name']}",
       ft=FT_TITLE, fg=F_NAVY, al=AC)
    ws.row_dimensions[row].height = 21.75
    row += 1

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    mw(ws, row, 1, row, N,
       f"DayOne CTP Building D  |  {data['for_desc']}",
       ft=FT_SUB, fg=F_BLUE, al=AC)
    ws.row_dimensions[row].height = 18
    row += 1

    # ── Rows 3-5: Info ────────────────────────────────────────────────────────
    def info_row(r, lbl_l, val_l, lbl_r, val_r):
        w(ws, r, 1, lbl_l, ft=FT_LABEL, fg=F_LBLUE, al=AL)
        mw(ws, r, 2, r, 5, val_l, ft=FT_BODY, fg=F_WHITE, al=AL)
        w(ws, r, 6, lbl_r, ft=FT_LABEL, fg=F_LBLUE, al=AL)
        mw(ws, r, 7, r, N, val_r, ft=FT_BODY, fg=F_WHITE, al=AL)
        ws.row_dimensions[r].height = 15.75

    inc = data['incomer']
    info_row(row, "Board Designation:", data['name'],
                  "Supply Source:", data['from_supply'])
    row += 1
    info_row(row, "Location:", "See Board Schedule",
                  "Incomer CB:",
                  f"{inc['rating']} {inc['poles']} {inc['type']}, {inc['brk']}")
    row += 1
    info_row(row, "Busbar Rating:", data['busbar'],
                  "Incomer Cable:", data['feed_size'])
    row += 1

    # ── Row 6: Spacer ─────────────────────────────────────────────────────────
    for c in range(1, N + 1):
        ws.cell(row=row, column=c).fill = F_WHITE
    ws.row_dimensions[row].height = 6
    row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION A — INCOMER
    # ══════════════════════════════════════════════════════════════════════════
    section_bar(ws, row, "SECTION A — INCOMER")
    row += 1
    col_headers(ws, row)
    row += 1

    r = row
    w(ws, r, 1,  "INCOMER",               ft=FT_BOLD,  fg=F_PEACH, al=AC)
    w(ws, r, 2,  data['from_supply'],     ft=FT_BOLD,  fg=F_PEACH, al=AL)
    w(ws, r, 3,  inc['rating'],           ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 4,  inc['type'],             ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 5,  inc['poles'],            ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 6,  inc['brk'],             ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 7,  None,                    ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 8,  None,                    ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 9,  None,                    ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 10, None,                    ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 11, data['feed_size'],       ft=FT_BODY,  fg=F_PEACH, al=AL)
    w(ws, r, 12, data['feed_route'],      ft=FT_BODY,  fg=F_PEACH, al=AL)
    w(ws, r, 13, None,                    ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws, r, 14, "ACTIVE",               ft=FT_BOLD,  fg=F_PEACH, al=AC)
    ws.row_dimensions[r].height = 27.75
    row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION B — OUTGOING CIRCUITS
    # ══════════════════════════════════════════════════════════════════════════
    section_bar(ws, row, "SECTION B — OUTGOING CIRCUITS")
    row += 1
    col_headers(ws, row)
    row += 1

    seen_groups = []
    for c in data['circuits']:
        if c['group'] not in seen_groups:
            seen_groups.append(c['group'])

    alt = 0
    for grp in seen_groups:
        grp_ckts = [c for c in data['circuits'] if c['group'] == grp]
        if not grp_ckts:
            continue

        # RCCB sub-group header row
        first = grp_ckts[0]
        if first['rccb_rating']:
            rccb_label = (f"{first['rccb_rating']} {first['rccb_poles']} "
                          f"{first['rccb_type']}  —  {first['desc']}")
        else:
            rccb_label = first['desc']
        mw(ws, row, 1, row, N,
           f"GROUP  {grp}   |   {rccb_label}",
           ft=FT_RCCB, fg=F_RCCB, al=AL)
        ws.row_dimensions[row].height = 15.75
        row += 1

        for c in grp_ckts:
            fg_row = F_GREEN if c['status'] == 'SPARE' \
                     else (F_LBLUE if alt % 2 == 0 else F_WHITE)
            circuit_row(ws, row,
                        c['circuit_no'], c['desc'],
                        c['cb_rating'], c['cb_type'], c['cb_poles'], c['cb_brk'],
                        c['kw'],
                        c['cable_size'], c['cable_route'],
                        '', c['status'], fg_row)
            alt += 1
            row += 1

        row += 1   # blank gap between groups

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION C — PANEL ACCESSORIES & METERING
    # ══════════════════════════════════════════════════════════════════════════
    acc = data['accessories']
    acc_rows = []
    if acc['spd']:
        acc_rows.append(("SPD", "Surge Protection Device (SPD)",
                         acc['spd'], "1 set",
                         f"with {acc['spd_mcb']} MCB" if acc['spd_mcb'] else ""))
    if acc['elr']:
        acc_rows.append(("ELR", "Earth Leakage Relay (ELR)", "—",
                         "1 no.", "Connected to ZCT"))
    if acc['zct']:
        acc_rows.append(("ZCT", "Zero Core Current Transformer (ZCT)", "—",
                         "1 no.", ""))
    if acc['pm']:
        acc_rows.append(("PM", "Power Meter (PM)", acc['pm'],
                         "1 no.", "BMS output"))
    if acc['bms']:
        acc_rows.append(("BMS", "BMS Interface / Transducer", acc['pm'],
                         "1 set", ""))
    if acc['lt_ctrl']:
        acc_rows.append(("LT", "Lighting Control Timer / Sensor",
                         acc['lt_ctrl'], "1 set", ""))
    if data['busbar']:
        acc_rows.append(("BUS", "Insulated CU Busbar", data['busbar'],
                         "1 set", ""))

    if acc_rows:
        section_bar(ws, row, "SECTION C — PANEL ACCESSORIES & METERING")
        row += 1
        w(ws,  row, 1, "Item",            ft=FT_HDR, fg=F_NAVY, al=AC)
        w(ws,  row, 2, "Description",     ft=FT_HDR, fg=F_NAVY, al=AL)
        w(ws,  row, 3, "Standard / Spec", ft=FT_HDR, fg=F_NAVY, al=AC)
        w(ws,  row, 4, "Qty",             ft=FT_HDR, fg=F_NAVY, al=AC)
        mw(ws, row, 5, row, N, "Remarks", ft=FT_HDR, fg=F_NAVY, al=AC)
        ws.row_dimensions[row].height = 18
        row += 1

        for i, (item, desc, spec, qty, rmk) in enumerate(acc_rows):
            fg = F_LBLUE if i % 2 == 0 else F_WHITE
            w(ws,  row, 1, item, ft=FT_BODY, fg=fg, al=AC)
            w(ws,  row, 2, desc, ft=FT_BODY, fg=fg, al=AL)
            w(ws,  row, 3, spec, ft=FT_BODY, fg=fg, al=AC)
            w(ws,  row, 4, qty,  ft=FT_BODY, fg=fg, al=AC)
            mw(ws, row, 5, row, N, rmk, ft=FT_BODY, fg=fg, al=AC)
            ws.row_dimensions[row].height = 18
            row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # NOTES
    # ══════════════════════════════════════════════════════════════════════════
    section_bar(ws, row, "NOTES")
    row += 1

    notes = [
        f"1.  Board designation: {data['name']}",
        f"2.  Incomer supply {data['from_supply']} via {data['feed_size']}"
        + (f" on {data['feed_route']}." if data['feed_route'] else "."),
        f"3.  Incomer CB: {inc['rating']} {inc['poles']} {inc['type']}, {inc['brk']}.",
        "4.  Installed capacity (kW) and demand factor to be verified by M&E Engineer.",
        "5.  ⚠️  QUERY RECOMMENDED: Confirm kW values and circuit descriptions against issued IFC drawings.",
        f"6.  Drawing ref: {data['name']} (DXF schematic). Verify before pricing.",
    ]
    note_fills = [F_LGRAY, F_WHITE, F_LGRAY, F_WHITE, F_YELL, F_LGRAY]
    for i, note in enumerate(notes):
        fg = note_fills[i] if i < len(note_fills) else F_WHITE
        mw(ws, row, 1, row, N, note, ft=FT_BODY, fg=fg, al=AL)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── Freeze & print ────────────────────────────────────────────────────────
    ws.freeze_panes = "A10"
    ws.print_title_rows = "1:9"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"Saved → {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_single_db.py <input.dxf>")
        sys.exit(1)

    dxf_path = sys.argv[1]
    _base    = os.path.dirname(os.path.abspath(__file__))
    stem     = os.path.splitext(os.path.basename(dxf_path))[0]
    out_path = os.path.join(_base, "output", f"{stem}_LoadSchedule.xlsx")

    print(f"Reading: {os.path.abspath(dxf_path)}")
    ents = load_ents(dxf_path)
    print(f"  {len(ents)} text entities loaded")

    data = extract_board(ents)
    print(f"  Board  : {data['name']}")
    print(f"  FOR    : {data['for_desc'][:60]}")
    print(f"  FROM   : {data['from_supply']}")
    inc = data['incomer']
    print(f"  Incomer: {inc['rating']} {inc['poles']} {inc['type']}, {inc['brk']}")

    gc = Counter(c['group'] for c in data['circuits'])
    for grp, cnt in sorted(gc.items()):
        print(f"    Group {grp}: {cnt} circuits")

    write_excel(data, out_path)


if __name__ == '__main__':
    main()
