"""
Parse DXF electrical load schedule (D1-MDB-COM-FOH1A-A) and export to Excel.
"""
import ezdxf
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_FILE = r"C:\Users\123\Desktop\D1-MDB-COM-FOH1A-A.dxf"
XLSX_FILE = r"C:\Users\123\Desktop\D1-MDB-COM-FOH1A-A_LoadTable.xlsx"


def clean_mtext(raw: str) -> str:
    """Strip MTEXT formatting codes."""
    text = raw
    # Remove formatting groups like {\fFont|...; text}
    text = re.sub(r'\{\\[^}]*\}', '', text)
    # Remove inline format codes like \pxqc; \A1; etc.
    text = re.sub(r'\\[A-Za-z][^;]*;', '', text)
    # Replace MTEXT paragraph break \P with newline
    text = text.replace(r'\P', '\n').replace(r'\p', '\n')
    # Remove remaining backslash commands
    text = re.sub(r'\\[A-Z]', '', text)
    # Unicode superscript fix
    text = text.replace('Â²', '²').replace('â‰¥', '≥').replace('ÂµS', 'µS')
    text = text.strip()
    return text


def extract_entities(doc):
    msp = doc.modelspace()
    entities = []
    for e in msp:
        if e.dxftype() == 'TEXT':
            x = e.dxf.insert.x
            y = e.dxf.insert.y
            txt = e.dxf.text
            entities.append({'type': 'TEXT', 'x': x, 'y': y, 'raw': txt,
                              'text': clean_mtext(txt)})
        elif e.dxftype() == 'MTEXT':
            x = e.dxf.insert.x
            y = e.dxf.insert.y
            txt = e.plain_text()
            entities.append({'type': 'MTEXT', 'x': x, 'y': y, 'raw': txt,
                              'text': txt.strip()})
    return entities


def group_by_y(entities, tolerance=150):
    """Group entities into rows by Y coordinate."""
    sorted_ents = sorted(entities, key=lambda e: -e['y'])
    rows = []
    for ent in sorted_ents:
        placed = False
        for row in rows:
            if abs(row['y_center'] - ent['y']) <= tolerance:
                row['items'].append(ent)
                # update center
                row['y_center'] = sum(i['y'] for i in row['items']) / len(row['items'])
                placed = True
                break
        if not placed:
            rows.append({'y_center': ent['y'], 'items': [ent]})
    # sort each row by X
    for row in rows:
        row['items'].sort(key=lambda e: e['x'])
    return rows


def main():
    print(f"Opening {DXF_FILE} ...")
    doc = ezdxf.readfile(DXF_FILE)
    entities = extract_entities(doc)
    print(f"Found {len(entities)} text entities")

    # Print all for inspection
    for e in sorted(entities, key=lambda x: -x['y']):
        print(f"  Y={e['y']:.0f}  X={e['x']:.0f}  [{e['type']}]  {repr(e['text'])}")

    rows = group_by_y(entities, tolerance=250)

    print(f"\nGrouped into {len(rows)} rows")
    for r in rows:
        texts = [it['text'] for it in r['items']]
        print(f"  Y≈{r['y_center']:.0f}: {texts}")

    build_excel(rows, entities)


def build_excel(rows, entities):
    # ─── Identify column positions ───────────────────────────────────────────
    # The leftmost column (X≈735) has row-header labels.
    # Each circuit occupies a band of X values.
    # We detect circuit X-positions from "CIRCUIT NO:" row values.
    circuit_row = None
    for row in rows:
        texts = [it['text'] for it in row['items']]
        if any('CIRCUIT NO' in t or 'P-1RYB' in t or 'RYB' in t for t in texts):
            if any('P-1RYB' in t or '2RYB' in t or '3RYB' in t for t in texts):
                circuit_row = row
                break

    # Collect all entities sorted
    all_sorted = sorted(entities, key=lambda e: -e['y'])

    # Determine distinct X column centres using clustering
    all_x = sorted(set(round(e['x'] / 500) * 500 for e in entities))

    # ─── Build structured table from known rows ───────────────────────────────
    # Find key rows by known text markers
    def find_row_by_text(marker):
        for r in rows:
            for it in r['items']:
                if marker.lower() in it['text'].lower():
                    return r
        return None

    # Key row lookups
    row_circuit   = find_row_by_text('CIRCUIT NO')
    row_room      = find_row_by_text('ROOM / EQUIPMENT')
    row_name      = find_row_by_text('NAME')
    row_installed = find_row_by_text('INSTALLATED')
    row_capacity  = find_row_by_text('CAPACITY')
    row_l1        = find_row_by_text('L1')
    row_cable     = find_row_by_text('CABLE AND CABLE')
    row_cable_type = find_row_by_text('XLPE')
    row_st_fs = None
    for r in rows:
        texts = [it['text'] for it in r['items']]
        if 'ST' in texts and 'FS' in texts:
            row_st_fs = r
            break

    # Find panel title row (top)
    row_title = find_row_by_text('D1-MDB-COM-FOH1A-A')
    row_source = find_row_by_text('FROM D1-SB')
    row_total = find_row_by_text('TOTAL: 1')
    row_for = find_row_by_text('FOR  COMMON')

    # Find circuit X positions (from circuit no row and sub-circuit rows)
    # Get all entities at Y near bottom (circuit numbers)
    circuit_ents = sorted(
        [e for e in entities if 200 < e['y'] < 600],
        key=lambda e: e['x']
    )

    # Gather circuits
    circuits = []
    for ent in circuit_ents:
        txt = ent['text']
        if re.search(r'RYB|SPARE', txt, re.I):
            circuits.append({'x': ent['x'], 'circuit_no': txt})

    if not circuits:
        # fallback: use all circuit-like texts
        for r in rows:
            for it in r['items']:
                if re.search(r'\dRYB|P-\dRYB', it['text']):
                    circuits.append({'x': it['x'], 'circuit_no': it['text']})

    circuits.sort(key=lambda c: c['x'])
    print(f"\nCircuits found: {[c['circuit_no'] for c in circuits]}")

    # Fetch text value at (x_ref ± tol, y_ref ± tol)
    def get_text_near(x_ref, y_ref, x_tol=1500, y_tol=250):
        matches = [e for e in entities
                   if abs(e['x'] - x_ref) <= x_tol and abs(e['y'] - y_ref) <= y_tol]
        if not matches:
            return ''
        matches.sort(key=lambda e: abs(e['x'] - x_ref) + abs(e['y'] - y_ref))
        return matches[0]['text']

    def get_row_texts(row):
        """Return all texts in a row sorted by X."""
        if row is None:
            return []
        return [it['text'] for it in sorted(row['items'], key=lambda i: i['x'])]

    # ─── Build Excel ──────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Schedule"

    # Styles
    title_font  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    header_font = Font(name='Arial', bold=True, size=10)
    body_font   = Font(name='Arial', size=9)
    bold_font   = Font(name='Arial', bold=True, size=9)
    small_font  = Font(name='Arial', size=8)

    blue_fill  = PatternFill('solid', fgColor='1F4E79')
    teal_fill  = PatternFill('solid', fgColor='1F7A8C')
    grey_fill  = PatternFill('solid', fgColor='D9D9D9')
    lblue_fill = PatternFill('solid', fgColor='BDD7EE')
    yellow_fill= PatternFill('solid', fgColor='FFFF99')
    green_fill = PatternFill('solid', fgColor='E2EFDA')
    orange_fill= PatternFill('solid', fgColor='FCE4D6')

    thin  = Side(style='thin',   color='000000')
    thick = Side(style='medium', color='000000')

    def border(l='thin', r='thin', t='thin', b='thin'):
        sides = {'thin': thin, 'thick': thick, None: Side(style=None)}
        return Border(
            left=sides.get(l, thin), right=sides.get(r, thin),
            top=sides.get(t, thin),  bottom=sides.get(b, thin)
        )

    def set_cell(ws, row, col, value, font=None, fill=None, align=None, bdr=None, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        if font:  c.font = font
        if fill:  c.fill = fill
        if align: c.alignment = align
        elif wrap:c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        if bdr:   c.border = bdr
        return c

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center', wrap_text=True)

    def merge_and_set(ws, r1, c1, r2, c2, value, font=None, fill=None, align=None):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=value)
        if font:  c.font = font
        if fill:  c.fill = fill
        c.alignment = align if align else center
        for row in range(r1, r2+1):
            for col in range(c1, c2+1):
                ws.cell(row=row, column=col).border = border()
        return c

    # ─── Collect all unique data ───────────────────────────────────────────────
    # Gather cable sizes at each circuit X
    # Find all MTEXT/TEXT entries and bin by circuit column
    def get_col_values(circuits, y_ref, y_tol=300):
        result = {}
        for c in circuits:
            val = get_text_near(c['x'], y_ref, x_tol=1800, y_tol=y_tol)
            result[c['circuit_no']] = val
        return result

    # Gather destination labels (circuit load names like D1-DB-...)
    # These appear as MTEXT entries near top area of circuit columns
    dest_y = None
    dest_ents = [e for e in entities if 'D1-DB' in e['text'] or 'D1-SB' in e['text']]
    print(f"\nDestination entities: {[(e['x'], e['y'], e['text']) for e in dest_ents]}")

    cb_ents = [e for e in entities if 'MCCB' in e['text'] or 'MCB' in e['text']]
    print(f"\nCB entities: {[(e['x'], e['y'], e['text']) for e in cb_ents]}")

    cable_ents = [e for e in entities if 'mm' in e['text'] or 'CU' in e['text']]
    print(f"\nCable entities: {[(e['x'], e['y'], e['text']) for e in cable_ents]}")

    spare_ents = [e for e in entities if 'SPARE' in e['text'].upper()]
    print(f"\nSpare entities: {[(e['x'], e['y'], e['text']) for e in spare_ents]}")

    # ──────────────────────────────────────────────────────────────────────────
    # ROW LAYOUT:
    # Row 1  : Title bar
    # Row 2  : Panel info (source, total)
    # Row 3  : For area
    # Row 4  : Spacer / main breaker header
    # Row 5  : Main breaker info
    # Row 6  : Bus bar
    # Row 7  : Instruments (PM, ELR, SPD, ZCT)
    # Row 8  : Phase (L1/L2/L3) columns header
    # Row 9-on: Per-circuit rows
    # ──────────────────────────────────────────────────────────────────────────

    # Derive per-circuit data
    # Distinct Y bands in the drawing
    ybands = sorted(set(round(e['y']/500)*500 for e in entities), reverse=True)

    # Find all entities grouped
    print("\nAll row groups:")
    for r in rows:
        texts = [it['text'] for it in r['items']]
        print(f"  Y≈{r['y_center']:.0f}: {texts}")

    # ─── Extract key info ─────────────────────────────────────────────────────
    panel_name = 'D1-MDB-COM-FOH1A-A'
    panel_for  = next((e['text'] for e in entities if 'COMMON AREA' in e['text']), '')
    panel_from = next((e['text'] for e in entities if 'FROM D1-SB' in e['text']), '')
    panel_total= next((e['text'] for e in entities if 'TOTAL' in e['text']), '')

    main_cb_text = next((e['text'] for e in entities
                         if '160A TPN' in e['text'] or '160AT' in e['text']), '')
    busbar_text  = next((e['text'] for e in entities if 'BUSBAR' in e['text']), '')
    incoming_cable = next((e['text'] for e in entities if '70mm' in e['text']), '')

    # Instruments rows
    instr_list = []
    for kw in ['PM', 'ELR', 'SPD', 'ZCT']:
        e = next((x for x in entities if x['text'].strip() == kw), None)
        if e:
            instr_list.append(f"{kw}  X={e['x']:.0f}")

    # Per-circuit data: collect destination name, CB, cable for each circuit column
    circ_data = []
    for c in circuits:
        cx = c['x']
        # Destination/Load name
        dest = get_text_near(cx, 8000, x_tol=2000, y_tol=2000)
        # CB rating near ST/FS row area
        cb = get_text_near(cx, 8700, x_tol=2000, y_tol=500)
        # Cable
        cable = get_text_near(cx, 5000, x_tol=2000, y_tol=3000)
        # Room/equipment
        room = get_text_near(cx, 2600, x_tol=2000, y_tol=300)
        # Name
        name = get_text_near(cx, 2350, x_tol=2000, y_tol=300)
        # Installed kW
        inst = get_text_near(cx, 1650, x_tol=2000, y_tol=300)
        # Capacity kW
        cap  = get_text_near(cx, 1350, x_tol=2000, y_tol=300)
        # ST/FS
        st  = get_text_near(cx, 8475, x_tol=2000, y_tol=200)

        circ_data.append({
            'circuit_no': c['circuit_no'],
            'dest': dest,
            'cb': cb,
            'cable': cable,
            'room': room,
            'name': name,
            'installed': inst,
            'capacity': cap,
            'st': st,
        })
        print(f"\nCircuit {c['circuit_no']}: dest={dest!r} cb={cb!r} cable={cable!r} room={room!r} name={name!r} inst={inst!r} cap={cap!r}")

    # ─── Write Excel ──────────────────────────────────────────────────────────
    # Column layout
    # Col A: Row label / description
    # Col B onwards: one per circuit

    # Number of circuits
    n_circ = len(circuits)
    n_cols = 1 + n_circ  # label col + circuit cols

    # Set column widths
    ws.column_dimensions['A'].width = 28
    for i in range(n_circ):
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = 22

    current_row = 1

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 30
    merge_and_set(ws, current_row, 1, current_row, n_cols,
                  f"DISTRIBUTION BOARD LOAD SCHEDULE",
                  font=Font(name='Arial', bold=True, size=14, color='FFFFFF'),
                  fill=blue_fill, align=center)
    current_row += 1

    # ── Row 2: Panel name ─────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 22
    merge_and_set(ws, current_row, 1, current_row, n_cols,
                  f"PANEL: {panel_name}",
                  font=Font(name='Arial', bold=True, size=12, color='FFFFFF'),
                  fill=teal_fill, align=center)
    current_row += 1

    # ── Row 3: Source / For ───────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 18
    half = max(1, n_cols // 2)
    merge_and_set(ws, current_row, 1, current_row, half,
                  f"SOURCE: {panel_from}",
                  font=bold_font, fill=lblue_fill, align=left)
    merge_and_set(ws, current_row, half+1, current_row, n_cols,
                  f"{panel_for}  |  {panel_total}",
                  font=bold_font, fill=lblue_fill, align=right)
    current_row += 1

    # ── Row 4: Incoming cable / Main breaker ──────────────────────────────────
    ws.row_dimensions[current_row].height = 36
    merge_and_set(ws, current_row, 1, current_row, 1,
                  "INCOMING CABLE", font=bold_font, fill=grey_fill, align=center)
    merge_and_set(ws, current_row, 2, current_row, n_cols,
                  incoming_cable,
                  font=body_font, fill=yellow_fill, align=center)
    current_row += 1

    ws.row_dimensions[current_row].height = 30
    merge_and_set(ws, current_row, 1, current_row, 1,
                  "MAIN CIRCUIT BREAKER", font=bold_font, fill=grey_fill, align=center)
    merge_and_set(ws, current_row, 2, current_row, n_cols,
                  main_cb_text,
                  font=body_font, fill=yellow_fill, align=center)
    current_row += 1

    ws.row_dimensions[current_row].height = 22
    merge_and_set(ws, current_row, 1, current_row, 1,
                  "BUS BAR", font=bold_font, fill=grey_fill, align=center)
    merge_and_set(ws, current_row, 2, current_row, n_cols,
                  busbar_text,
                  font=body_font, fill=yellow_fill, align=center)
    current_row += 1

    # ── Row 7: Instruments ────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 18
    pm_ents  = [e for e in entities if e['text'].strip() == 'PM']
    elr_ents = [e for e in entities if e['text'].strip() == 'ELR']
    spd_ents = [e for e in entities if 'SPD' in e['text']]
    zct_ents = [e for e in entities if 'ZCT' in e['text']]
    instruments_str = '  |  '.join(filter(None, [
        'PM' if pm_ents else '',
        'ELR' if elr_ents else '',
        'SPD' if spd_ents else '',
        'ZCT' if zct_ents else '',
    ]))
    merge_and_set(ws, current_row, 1, current_row, 1,
                  "INSTRUMENTATION", font=bold_font, fill=grey_fill, align=center)
    merge_and_set(ws, current_row, 2, current_row, n_cols,
                  instruments_str,
                  font=body_font, fill=green_fill, align=center)
    current_row += 1

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 40
    headers = ['DESCRIPTION / ITEM'] + [c['circuit_no'] for c in circuits]
    sub_headers = [''] + ['' for _ in circuits]

    for col_i, hdr in enumerate(headers):
        c = ws.cell(row=current_row, column=col_i+1, value=hdr)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = blue_fill
        c.alignment = center
        c.border = border('thick','thick','thick','thick')
    current_row += 1

    # ── Circuit data rows ─────────────────────────────────────────────────────
    row_specs = [
        ('LOAD / DESTINATION',    'dest',      orange_fill),
        ('CIRCUIT BREAKER',       'cb',         grey_fill),
        ('CABLE SIZE & TYPE',     'cable',      lblue_fill),
        ('ROOM / EQUIPMENT',      'room',       green_fill),
        ('LOAD NAME',             'name',       green_fill),
        ('INSTALLED CAPACITY (kW)','installed', yellow_fill),
        ('MAX DEMAND (kW)',        'capacity',  yellow_fill),
        ('START / RUN TYPE',      'st',         grey_fill),
    ]

    for label, key, fill in row_specs:
        ws.row_dimensions[current_row].height = 36
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = bold_font
        c.fill = grey_fill
        c.alignment = left
        c.border = border()
        for col_i, cd in enumerate(circ_data):
            val = cd.get(key, '')
            cell = ws.cell(row=current_row, column=col_i+2, value=val)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='center',
                                        horizontal='center')
            cell.border = border()
        current_row += 1

    # ── Phase loading row (L1/L2/L3) ─────────────────────────────────────────
    for phase in ['L1', 'L2', 'L3']:
        ws.row_dimensions[current_row].height = 22
        c = ws.cell(row=current_row, column=1, value=f"PHASE {phase} LOADING")
        c.font = bold_font; c.fill = grey_fill
        c.alignment = left; c.border = border()
        for col_i in range(n_circ):
            cell = ws.cell(row=current_row, column=col_i+2, value='')
            cell.fill = lblue_fill; cell.border = border()
        current_row += 1

    # ── Total row ─────────────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 22
    merge_and_set(ws, current_row, 1, current_row, n_cols,
                  f"TOTAL CIRCUITS: {n_circ}",
                  font=Font(name='Arial', bold=True, size=10, color='FFFFFF'),
                  fill=teal_fill, align=right)
    current_row += 1

    # ── Notes section ─────────────────────────────────────────────────────────
    current_row += 1
    ws.row_dimensions[current_row].height = 18
    c = ws.cell(row=current_row, column=1, value="NOTES:")
    c.font = bold_font; c.fill = grey_fill
    c.alignment = left
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=n_cols)
    current_row += 1

    # SPD info
    spd_text = next((e['text'] for e in entities if 'TYPE 2' in e['text'] or 'IEC61643' in e['text']), '')
    if spd_text:
        spd_clean = re.sub(r'\\[^\\]+;|[{}]', '', spd_text).replace(r'\H', '').strip()
        merge_and_set(ws, current_row, 1, current_row, n_cols,
                      f"SPD: {spd_clean}",
                      font=body_font, fill=None, align=left)
        current_row += 1

    # Freeze top rows
    ws.freeze_panes = 'B9'

    wb.save(XLSX_FILE)
    print(f"\nExcel saved to: {XLSX_FILE}")


if __name__ == '__main__':
    main()
