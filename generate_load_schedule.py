"""
Generate D1-MDB-COM-FOH1A-A_LoadSchedule_Generated.xlsx
Replicates the exact format of D1-MDB-COM-FOH1A-A_LoadSchedule.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\123\Desktop\D1-MDB-COM-FOH1A-A_LoadSchedule_Generated.xlsx"

# ── Colours (RGB without alpha) ───────────────────────────────────────────────
def fill(hex6):
    return PatternFill("solid", fgColor="FF" + hex6.upper())

F_NAVY   = fill("1F3864")   # title / column header bg
F_BLUE   = fill("2E75B6")   # section bar bg
F_LBLUE  = fill("D6E4F0")   # info-label cells & odd active rows
F_WHITE  = fill("FFFFFF")   # even active rows / value cells
F_PEACH  = fill("FCE4D6")   # INCOMER row
F_GREEN  = fill("E2EFDA")   # SPARE rows
F_LGRAY  = fill("F2F2F2")   # alternating notes rows
F_YELLOW = fill("FFF2CC")   # warning note

# ── Fonts ─────────────────────────────────────────────────────────────────────
def font(bold=False, size=9, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size,
                color="FF" + color.upper(), italic=italic)

FT_TITLE   = font(bold=True,  size=13, color="FFFFFF")
FT_SUBTITLE= font(bold=False, size=9,  color="FFFFFF")
FT_HDR     = font(bold=True,  size=9,  color="FFFFFF")
FT_LABEL   = font(bold=True,  size=9,  color="000000")
FT_BODY    = font(bold=False, size=9,  color="000000")
FT_BOLD    = font(bold=True,  size=9,  color="000000")
FT_GREY    = font(bold=False, size=9,  color="7F7F7F")
FT_GREYBOLD= font(bold=True,  size=9,  color="7F7F7F")

# ── Alignments ────────────────────────────────────────────────────────────────
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left",   vertical="center", wrap_text=True)
AR = Alignment(horizontal="right",  vertical="center", wrap_text=True)

# ── Border ────────────────────────────────────────────────────────────────────
_thin = Side(style="thin", color="FF000000")
BD    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# ── Helpers ───────────────────────────────────────────────────────────────────
def w(ws, r, c, val, ft=FT_BODY, fg=F_WHITE, al=AC):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = ft
    cell.fill      = fg
    cell.alignment = al
    cell.border    = BD
    return cell

def mw(ws, r1, c1, r2, c2, val, ft=FT_BODY, fg=F_WHITE, al=AC):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=val)
    cell.font      = ft
    cell.fill      = fg
    cell.alignment = al
    cell.border    = BD
    # apply border to all merged cells
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(row=rr, column=cc).border = BD
    return cell

N = 14   # total columns A–N

def section_bar(ws, r, text):
    mw(ws, r, 1, r, N, text, ft=FT_HDR, fg=F_BLUE, al=AC)

def col_headers(ws, r, labels):
    for i, lbl in enumerate(labels, 1):
        w(ws, r, i, lbl, ft=FT_HDR, fg=F_NAVY, al=AC)

# ── Column headers shared between Section A & B ───────────────────────────────
HDR = [
    "Circuit No.",
    "Description / Load Name",
    "CB Rating (A)",
    "CB Type",
    "No. of Poles",
    "Breaking Cap.",
    "Connected Load (kW)",
    "Demand Factor",
    "Max Demand (kW)",
    "Current (A)",
    "Cable Size",
    "Cable Route",
    "Remarks",
    "Status",
]

# ── Data ──────────────────────────────────────────────────────────────────────
ACTIVE_CIRCUITS = [
    {
        "no":    "P-1RYB",
        "desc":  "D1-DB-MMR3S1B-A\n(MMR / Server Room Sub-Board)",
        "cb_a":  "63A",  "cb_t": "MCCB", "poles": "3P+N", "brk": "36kA",
        "cable": "4×1C 16mm² LSZH CU\n+1C 16mm² CU CPC",
        "route": "Cable Trunking / Conduit",
        "row":   12,
    },
    {
        "no":    "2RYB",
        "desc":  "D1-DB-COM-FOH1A-A\n(Common Area / FOH Sub-Distribution Board)",
        "cb_a":  "63A",  "cb_t": "MCCB", "poles": "3P+N", "brk": "36kA",
        "cable": "4×1C 16mm² LSZH CU\n+1C 16mm² CU CPC",
        "route": "Cable Trunking / Conduit",
        "row":   13,
    },
]
SPARE_CIRCUITS = ["3RYB", "4RYB", "5RYB", "6RYB", "7RYB", "8RYB"]

ACCESSORIES = [
    ("1", "Surge Protection Device (SPD)",    "Type 2, In≥40kA, IEC 61643",  "1 set",  "8/20µs, with 6A MCB"),
    ("2", "Earth Leakage Relay (ELR)",         "—",                           "1 no.",  "Connected to ZCT"),
    ("3", "Zero Core Current Transformer (ZCT)","—",                          "1 no.",  ""),
    ("4", "Power Meter (PM)",                  "Class 0.5, BMS output",       "1 no.",  "Connected to BMS"),
    ("5", "BMS Interface / Transducer",        "Class 0.5",                   "1 set",  ""),
    ("6", "160A 4P Insulated CU Busbar",       "160A rated, 4 pole",          "1 set",  ""),
]

NOTES = [
    "1.  Board designation: D1-MDB-COM-FOH1A-A — FOR COMMON AREA (1 NO.)",
    "2.  Incomer supply from D1-SB-COM-FOH-BS1-A via 4×1C 70mm² XLPE/LSZH CU + 1C 35mm² CU CPC on cable ladder/tray.",
    "3.  All outgoing circuits rated 63A TPN MCCB 36kA. Feeder cables: 4×1C 16mm² LSZH CU + 1C 16mm² CU CPC on cable trunking/conduit.",
    "4.  FS = Fuse Switch; ST = Isolator Switch (shown on drawing per circuit).",
    "5.  Installed capacity (kW) and demand factor to be populated by M&E Engineer when load data is available.",
    "6.  ⚠️  QUERY RECOMMENDED: Room/Equipment names and installed capacity (kW) not embedded in DXF — confirm with panel schedule or M&E Engineer.",
    "7.  Drawing ref: D1-MDB-COM-FOH1A-A (DXF schematic). Verify against issued IFC drawings before pricing.",
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Schedule"
    ws.sheet_properties.tabColor = "2E75B6"

    # ── Column widths ──────────────────────────────────────────────────────────
    col_w = {
        "A": 16.86, "B": 38.0,  "C": 22.86, "D": 14.0,
        "E": 9.0,   "F": 14.0,  "G": 11.0,  "H": 11.0,
        "I": 12.0,  "J": 10.0,  "K": 28.0,  "L": 22.0,
        "M": 12.0,  "N": 9.0,
    }
    for col, width in col_w.items():
        ws.column_dimensions[col].width = width

    # ── Row heights ────────────────────────────────────────────────────────────
    row_h = {
        1: 21.75, 2: 18.0,  3: 15.75, 4: 15.75, 5: 15.75,
        6: 6.0,   7: 15.75, 8: 27.75, 9: 24.0,  10: 33.75,
        11: 27.75,
    }
    for r in range(12, 20):   row_h[r] = 27.75
    row_h[20] = 15.75
    row_h[21] = 19.5
    for r in range(22, 28):   row_h[r] = 18.0
    row_h[28] = 15.75
    for r in range(29, 36):   row_h[r] = 18.0
    for r, h in row_h.items():
        ws.row_dimensions[r].height = h

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 1 — Title
    # ══════════════════════════════════════════════════════════════════════════
    mw(ws, 1, 1, 1, N,
       "LOAD SCHEDULE — D1-MDB-COM-FOH1A-A",
       ft=FT_TITLE, fg=F_NAVY, al=AC)

    # ROW 2 — Subtitle
    mw(ws, 2, 1, 2, N,
       "DayOne CTP Building D | Common Area Main Distribution Board",
       ft=FT_SUBTITLE, fg=F_BLUE, al=AC)

    # ROW 3 — Board / Supply Source
    w(ws,  3, 1, "Board Designation:", ft=FT_LABEL, fg=F_LBLUE, al=AL)
    mw(ws, 3, 2, 3, 5, "D1-MDB-COM-FOH1A-A",      ft=FT_BODY, fg=F_WHITE, al=AL)
    w(ws,  3, 6, "Supply Source:",     ft=FT_LABEL, fg=F_LBLUE, al=AL)
    mw(ws, 3, 7, 3, N, "D1-SB-COM-FOH-BS1-A",     ft=FT_BODY, fg=F_WHITE, al=AL)

    # ROW 4 — Location / Incomer CB
    w(ws,  4, 1, "Location:",          ft=FT_LABEL, fg=F_LBLUE, al=AL)
    mw(ws, 4, 2, 4, 5, "Common Area / FOH",        ft=FT_BODY, fg=F_WHITE, al=AL)
    w(ws,  4, 6, "Incomer CB:",        ft=FT_LABEL, fg=F_LBLUE, al=AL)
    mw(ws, 4, 7, 4, N, "160AT/160AF TPN MCCB, 36kA (L,S,I)",
       ft=FT_BODY, fg=F_WHITE, al=AL)

    # ROW 5 — Busbar / Incomer Cable
    w(ws,  5, 1, "Busbar Rating:",     ft=FT_LABEL, fg=F_LBLUE, al=AL)
    mw(ws, 5, 2, 5, 5, "160A, 4P Insulated CU Busbar",
       ft=FT_BODY, fg=F_WHITE, al=AL)
    w(ws,  5, 6, "Incomer Cable:",     ft=FT_LABEL, fg=F_LBLUE, al=AL)
    mw(ws, 5, 7, 5, N,
       "4×1C 70mm² XLPE/LSZH CU + 1C 35mm² CU CPC",
       ft=FT_BODY, fg=F_WHITE, al=AL)

    # ROW 6 — spacer (fill cells so border is consistent)
    for c in range(1, N + 1):
        ws.cell(row=6, column=c).fill = F_WHITE

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION A — INCOMER
    # ══════════════════════════════════════════════════════════════════════════
    section_bar(ws, 7, "SECTION A — INCOMER")
    col_headers(ws, 8, HDR)

    # INCOMER data row (row 9)
    r = 9
    w(ws,  r, 1,  "INCOMER",                   ft=FT_BOLD,  fg=F_PEACH, al=AC)
    w(ws,  r, 2,  "Supply from D1-SB-COM-FOH-BS1-A",
                                                ft=FT_BOLD,  fg=F_PEACH, al=AL)
    w(ws,  r, 3,  "160A",                       ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 4,  "MCCB (L,S,I)",              ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 5,  "3P+N",                       ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 6,  "36kA",                       ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 7,  "-",                          ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 8,  "-",                          ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 9,  "-",                          ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 10, "-",                          ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 11, "4×1C 70mm² XLPE/LSZH CU\n+1C 35mm² CU CPC",
                                                ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 12, "Cable Ladder / Tray",        ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 13, "160AT/160AF",                ft=FT_BODY,  fg=F_PEACH, al=AC)
    w(ws,  r, 14, "ACTIVE",                     ft=FT_BODY,  fg=F_PEACH, al=AC)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION B — OUTGOING CIRCUITS
    # ══════════════════════════════════════════════════════════════════════════
    section_bar(ws, 10, "SECTION B — OUTGOING CIRCUITS (FEEDERS TO SUB-BOARDS)")
    col_headers(ws, 11, HDR)

    # Active circuits
    for i, circ in enumerate(ACTIVE_CIRCUITS):
        r    = circ["row"]
        fg_r = F_LBLUE if i % 2 == 0 else F_WHITE   # alternate row colour
        w(ws, r, 1,  circ["no"],    ft=FT_BOLD,  fg=fg_r, al=AC)
        w(ws, r, 2,  circ["desc"],  ft=FT_BODY,  fg=fg_r, al=AL)
        w(ws, r, 3,  circ["cb_a"],  ft=FT_BODY,  fg=fg_r, al=AC)
        w(ws, r, 4,  circ["cb_t"],  ft=FT_BODY,  fg=fg_r, al=AC)
        w(ws, r, 5,  circ["poles"], ft=FT_BODY,  fg=fg_r, al=AC)
        w(ws, r, 6,  circ["brk"],   ft=FT_BODY,  fg=fg_r, al=AC)
        w(ws, r, 7,  None,          ft=FT_BODY,  fg=fg_r, al=AC)   # Connected Load
        w(ws, r, 8,  None,          ft=FT_BODY,  fg=fg_r, al=AC)   # Demand Factor
        # Max Demand formula  =G*H
        cell_md = ws.cell(row=r, column=9)
        cell_md.value     = f"=G{r}*H{r}"
        cell_md.font      = FT_BODY
        cell_md.fill      = fg_r
        cell_md.alignment = AC
        cell_md.border    = BD
        w(ws, r, 10, None,          ft=FT_BODY,  fg=fg_r, al=AC)   # Current
        w(ws, r, 11, circ["cable"], ft=FT_BODY,  fg=fg_r, al=AL)
        w(ws, r, 12, circ["route"], ft=FT_BODY,  fg=fg_r, al=AL)
        w(ws, r, 13, None,          ft=FT_BODY,  fg=fg_r, al=AC)   # Remarks
        w(ws, r, 14, "ACTIVE",      ft=FT_BODY,  fg=fg_r, al=AC)

    # Spare circuits  (rows 14–19)
    for i, cno in enumerate(SPARE_CIRCUITS):
        r = 14 + i
        w(ws, r, 1,  cno,          ft=FT_GREYBOLD, fg=F_GREEN, al=AC)
        w(ws, r, 2,  "SPARE",      ft=FT_GREY,     fg=F_GREEN, al=AL)
        w(ws, r, 3,  "63A",        ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 4,  "MCCB",       ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 5,  "3P+N",       ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 6,  "36kA",       ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 7,  "-",          ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 8,  "-",          ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 9,  "-",          ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 10, "-",          ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 11, "-",          ft=FT_GREY,     fg=F_GREEN, al=AL)
        w(ws, r, 12, "-",          ft=FT_GREY,     fg=F_GREEN, al=AL)
        w(ws, r, 13, "Spare Way",  ft=FT_GREY,     fg=F_GREEN, al=AC)
        w(ws, r, 14, "SPARE",      ft=FT_GREY,     fg=F_GREEN, al=AC)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION C — PANEL ACCESSORIES & METERING
    # ══════════════════════════════════════════════════════════════════════════
    section_bar(ws, 20, "SECTION C — PANEL ACCESSORIES & METERING")

    # Section C headers  (row 21)
    w(ws,  21, 1, "Item",             ft=FT_HDR, fg=F_NAVY, al=AC)
    w(ws,  21, 2, "Description",      ft=FT_HDR, fg=F_NAVY, al=AC)
    w(ws,  21, 3, "Standard / Spec",  ft=FT_HDR, fg=F_NAVY, al=AC)
    w(ws,  21, 4, "Qty",              ft=FT_HDR, fg=F_NAVY, al=AC)
    mw(ws, 21, 5, 21, N, "Remarks",  ft=FT_HDR, fg=F_NAVY, al=AC)

    for i, (item, desc, spec, qty, remarks) in enumerate(ACCESSORIES):
        r   = 22 + i
        fg_r = F_LBLUE if i % 2 == 0 else F_WHITE
        w(ws,  r, 1, item,    ft=FT_BODY, fg=fg_r, al=AC)
        w(ws,  r, 2, desc,    ft=FT_BODY, fg=fg_r, al=AL)
        w(ws,  r, 3, spec,    ft=FT_BODY, fg=fg_r, al=AC)
        w(ws,  r, 4, qty,     ft=FT_BODY, fg=fg_r, al=AC)
        mw(ws, r, 5, r, N, remarks, ft=FT_BODY, fg=fg_r, al=AC)

    # ══════════════════════════════════════════════════════════════════════════
    # NOTES
    # ══════════════════════════════════════════════════════════════════════════
    section_bar(ws, 28, "NOTES")

    note_fills = [F_LGRAY, F_WHITE, F_LGRAY, F_WHITE, F_LGRAY, F_YELLOW, F_LGRAY]
    for i, note in enumerate(NOTES):
        r    = 29 + i
        fg_n = note_fills[i] if i < len(note_fills) else F_WHITE
        mw(ws, r, 1, r, N, note, ft=FT_BODY, fg=fg_n, al=AL)

    # ══════════════════════════════════════════════════════════════════════════
    # FREEZE PANES & PRINT SETTINGS
    # ══════════════════════════════════════════════════════════════════════════
    ws.freeze_panes = "A12"
    ws.print_title_rows = "1:11"
    ws.page_setup.orientation     = "landscape"
    ws.page_setup.fitToPage       = True
    ws.page_setup.fitToWidth      = 1
    ws.page_setup.fitToHeight     = 0

    wb.save(OUT)
    print(f"Saved → {OUT}")


if __name__ == "__main__":
    build()
